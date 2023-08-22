# -*- coding: utf-8 -*-
import openpyxl as xl
import re
from Dengy import Ui_Dengy  # Это наш конвертированный файл дизайна
from PySide6 import QtWidgets, QtCore
from PySide6.QtWidgets import QProxyStyle, QStyle, QMessageBox, QWidget, QLabel, QVBoxLayout, QStyledItemDelegate
from PySide6.QtCore import Qt
from  PySide6.QtGui import QFont

class MyProxyStyle(QProxyStyle):
    def pixelMetric(self, QStyle_PixelMetric, option=None, widget=None):

        if QStyle_PixelMetric == QStyle.PM_SmallIconSize:
            return 24
        else:
            return QProxyStyle.pixelMetric(self, QStyle_PixelMetric, option, widget)

class DengyApp(QtWidgets.QMainWindow, Ui_Dengy):
    def __init__(self):
        super().__init__()
        self.file = None
        self.ui = Ui_Dengy()
        self.ui.setupUi(self)
        self.ui.pushButton_open_ved.clicked.connect(self.open_vedomost)
        self.ui.pushButton_open_reestr.clicked.connect(self.open_reestr)
        self.ui.pushButton_vnesti.clicked.connect(self.runing)

        self.ui.pushButton_open_ved_3.clicked.connect(self.open_vedomost_2)
        self.ui.pushButton_open_reestr1_3.clicked.connect(self.open_reestr_2)
        self.ui.pushButton_vnesti_3.clicked.connect(self.runing_2)

        self.ui.tableWidget_kassa.setColumnWidth(0,220)
        self.ui.tableWidget_kassa.setColumnWidth(1,200)
        self.ui.tableWidget_kassa.setColumnWidth(2,110)
        self.ui.pushButton_save_exel.clicked.connect(self.save)


    def save(self):
        pass
        # data = [['Абидина Анастасия Михайловна', 'Ведущий специалист отдела инфраструктуры', '67101.2'], ['Бронников Дмитрий Александрович', 'ведущий специалист отдела ГО и ЧС', '15976.86'], ['Воскобойник Александр Павлович', 'Ведущий специалист отдела по делам молодежной политики, культуры, спорта и туризма', '28757.94'], ['Кица Сергей Николаевич', 'Охраник', '18644.29'], ['Коваленко Алексей Сергеевич', 'Специалист 1 категории сектора гуманитарной помощи', '39961.4'], ['Козаков Александр Николаевич', 'Главный специалист хозяйственного отдела', '14739.6'], ['Кушнаренко Андрей Валентинович', 'Главный специалист инфраструктуры', '77380.4'], ['Леонов Иван Юрьевич', 'Водитель', '27671'], ['Луганская -Кузина Марина Александровна', 'Отдел по социальным вопросам.Главный специалист', '-12109.88'], ['Питомец Владислав Иванович', 'Водитель', '25035.86'], ['Саркисян Ирина Рафаиловна', 'Специалист 1 категории сектора гум помощи', '59941.19'], ['Терехов Алексей Анатольевич', 'Охранник', '26102.4'], ['Шевцова Ольга Анатольевна', 'Специалист 1 категории сектора гум помощи', '19980.06']]
        #
        # for row in range(len(data)):
        #     self.ui.tableWidget_kassa.setRowCount(len(data))
        #     for column in range(len(data[row])):
        #         item = QtWidgets.QTableWidgetItem()
        #         item.setText(data[row][column])
        #         self.ui.tableWidget_kassa.setItem(row, column, item)

    def open_vedomost(self):
        try:
            file = QtWidgets.QFileDialog.getOpenFileName(self, "Выбрать ведомость", " ", "All (*);;Excel (*.xlsx)",
                                                    "Excel (*.xlsx)")
            fileName = file[0]
            self.file = fileName
            self.ui.textEdit.setText(self.file)
        except AttributeError:
            self.ui.label_info.setText('Выберите файл для анализа.')

    def open_reestr(self):
        try:
            file = QtWidgets.QFileDialog.getOpenFileName(self, "Выбрать файл реестра", " ", "All (*);;Excel (*.xlsx)",
                                                    "Excel (*.xlsx)")
            fileName = file[0]
            self.file_reestr = fileName
            self.ui.textEdit_2.setText(self.file_reestr)
        except AttributeError:
            self.ui.label_info.setText('Выберите файл для анализа.')

    # Выбираем данные из ведомости и переносим в реестр
    def runing(self):
        self.ui.label_info.clear()
        column_start = self.ui.spinBox_column_start.value()
        column_start_2 = self.ui.spinBox_column_start_2.value()
        row_start = self.ui.spinBox_row_start.value()
        row_finish = self.ui.spinBox_row_finish.value()
        name_book = self.ui.lineEdit_book.text()

        if column_start == 0:
            self.ui.label_info.setText('Введите № начальной колонки для № счетов.')
        # if column_start_2 == 0:
        #     self.ui.label_info.setText('Введите № начальной колонки для сумм.')
        # elif row_start == 0:
        #     self.ui.label_info.setText('Введите № начальной строки для расчёта.')
        # elif row_finish == 0:
        #     self.ui.label_info.setText('Введите № конечной строки для расчёта.')
        # elif name_book == '':
        #     self.ui.label_info.setText('Введите имя книги в ведомости.')
        else:
            try:
                wb = xl.load_workbook(self.file, keep_vba=True, data_only=True)
                wb.active = wb[name_book]
                ws = wb.active

                scor_coord = []

                for col_cells in ws.iter_cols(min_row=row_start, max_row=row_finish, min_col=column_start,
                                              max_col=column_start_2):
                    for cell in col_cells:
                        if cell.value == '0':
                            summ = ws.cell(row=cell.row, column=column_start_2).value
                            scor_coord.append([cell.value, cell.coordinate, cell.row, round(summ, 2)])
                        # score_all.append(cell.value)
                        # print('%s: cell.value=%s' % (cell, cell.value))
                # ИЮЛЬ2023 Sheet0

                scor_coord_new = []
                itog = 0
                for coord in scor_coord:
                    itog += coord[3]
                    scor_coord_new.append([coord[0], 'C' + coord[1][1:], 'D' + coord[1][1:], coord[2], coord[3]])

                # print(scor_coord_new)
                data = []
                for part in scor_coord_new:
                    data.append([ws[part[1]].value, ws[part[2]].value, str(part[4])])
                    # part.append(ws[part[1]].value)
                    # print(ws.cell(row=part[2], column=25).value)

                print(data)

                for row in range(len(data)):
                    self.ui.tableWidget_kassa.setRowCount(len(data)+1)
                    for column in range(len(data[row])):
                        item = QtWidgets.QTableWidgetItem()
                        item.setText(data[row][column])
                        item.setFont(QFont('Times', 10))
                        self.ui.tableWidget_kassa.setItem(row, column, item)
                item2 = QtWidgets.QTableWidgetItem('ИТОГО:')
                item2.setFont(QFont('Times', 11, QFont.Black))
                item2.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter | Qt.AlignRight)
                self.ui.tableWidget_kassa.setItem(row+1, column-1, item2)
                item3 = QtWidgets.QTableWidgetItem(str(itog))
                item3.setFont(QFont('Times', 11, QFont.Black))
                self.ui.tableWidget_kassa.setItem(row+1, column, item3)


                # Выбираем все № счетов
                score_all = []
                summ_all = []
                for col_cells in ws.iter_cols(min_row=row_start, max_row=row_finish, min_col=column_start,
                                                     max_col=column_start):
                    for cell in col_cells:
                        score_all.append(cell.value)
                        # print('%s: cell.value=%s' % (cell, cell.value))
                # print('Проверка 1: ', score_all)
                for col_cells in ws.iter_cols(min_row=row_start, max_row=row_finish, min_col=column_start_2,
                                                     max_col=column_start_2):
                    for cell in col_cells:
                        rd = round(cell.value,2)
                        summ_all.append(rd)
                        # print('%s: cell.value=%s' % (cell, cell.value))
                # print('Проверка 1: ', summ_all)

                i = 0
                new_bill = []
                for score in score_all:
                    new_bill.append([score, summ_all[i]])
                    i+=1

                # print(score_all)

                # Находим дубликаты в списке
                dup = [x for i, x in enumerate(score_all) if i != score_all.index(x) and len(score_all[i])>15]
                # print(dup)  # [1, 5, 1]

                dup_new = []
                for i in dup:
                   dup_new.append([i,777])

                # print(new_bill)
                for t in dup_new:
                    for j in new_bill:
                        if j[0] == t[0] and t[1] == 777:
                            t[1] = j[1]
                        elif j[0] == t[0] and t[1] != 777:
                            t[1] = round(t[1] + j[1], 2)

                for k in dup:
                    for j in new_bill:
                        if k == j[0]:
                            new_bill.pop(new_bill.index(j))
                for k in dup:
                    for j in new_bill:
                        if k == j[0]:
                            new_bill.pop(new_bill.index(j))

                for k in dup_new:
                    for s in new_bill:
                        if k[0] == s[0]:
                            pass
                        else:
                            new_bill.append(k)
                        break
                wb.close()
                self.ui.label_info.setText('Данные успешно внесены!')
            except:
                self.ui.label_info.setText('Проверте веденные данные по ведомости.')

            # self.ui.label_info_4.clear()
            # column_start_reestr = self.ui.spinBox_column_start_11.value()
            # column_start_reestr_2 = self.ui.spinBox_column_start_12.value()
            # row_start_reestr = self.ui.spinBox_row_start_6.value()
            # row_finish_reestr = self.ui.spinBox_row_finish_6.value()
            # name_book_reestr = self.ui.lineEdit_book_3.text()
            #
            # if column_start_reestr == 0:
            #     self.ui.label_info_4.setText('Введите № начальной колонки для № счетов.')
            # if column_start_reestr_2 == 0:
            #     self.ui.label_info_4.setText('Введите № начальной колонки для сумм.')
            # elif row_start_reestr == 0:
            #     self.ui.label_info_4.setText('Введите № начальной строки для расчёта.')
            # elif row_finish_reestr == 0:
            #     self.ui.label_info_4.setText('Введите № конечной строки для расчёта.')
            # elif name_book_reestr == '':
            #     self.ui.label_info_4.setText('Введите имя книги в реестре.')
            # else:
            #     try:
            #         wb1 = xl.load_workbook(self.file_reestr)
            #         # print(wb1.sheetnames)
            #         wb1.active = wb1[name_book_reestr]
            #         ws2 = wb1.active
            #         # # Выбираем все № счетов
            #
            #         score = []
            #         coord = []
            #
            #         for col_cells in ws2.iter_cols(min_row=row_start_reestr, max_row=row_finish_reestr,
            #                                        min_col=column_start_reestr, max_col=column_start_reestr):
            #             for cell in col_cells:
            #                 score.append([cell.value])
            #
            #         for col_cells in ws2.iter_cols(min_row=row_start_reestr, max_row=row_finish_reestr,
            #                                                min_col=column_start_reestr_2, max_col=column_start_reestr_2):
            #             for cell in col_cells:
            #                 coord.append(cell.coordinate)
            #
            #         reestr1 = []
            #         t = 0
            #         for i in score:
            #             reestr1.append([i[0], coord[t]])
            #             t += 1
            #
            #
            #         for re in reestr1:
            #             for d in new_bill:
            #                 if re[0] == d[0]:
            #                     # print(d[1])
            #                     # print(re)
            #                     re.append(d[1])
            #         # print(reestr1)
            #
            #         for i in reestr1:
            #             try:
            #                 ch = str(i[2])
            #                 zam = ch.replace(",", ".")
            #                 ws2[i[1]] = zam
            #             except IndexError:
            #                 pass
            #
            #         wb1.save(self.file_reestr)
            #         self.ui.label_info_4.setText('Данные успешно внесены!')
            #     except:
            #         self.ui.label_info_4.setText('Проверте веденные данные веденные по реестру.')
    def open_vedomost_2(self):
        try:
            file = QtWidgets.QFileDialog.getOpenFileName(self, "Выбрать ведомость", " ", "All (*);;Excel (*.xlsx)",
                                                    "Excel (*.xlsx)")
            fileName = file[0]
            self.file_ved = fileName
            self.ui.textEdit_5.setText(self.file_ved)
        except AttributeError:
            self.ui.label_info_2.setText('Выберите файл для анализа.')

    def open_reestr_2(self):
        try:
            file = QtWidgets.QFileDialog.getOpenFileName(self, "Выбрать файл реестра", " ", "All (*);;Excel (*.xlsx)",
                                                    "Excel (*.xlsx)")
            fileName = file[0]
            self.file_reestr_1 = fileName
            self.ui.textEdit_6.setText(self.file_reestr_1)
        except AttributeError:
            self.ui.label_info_2.setText('Выберите файл для анализа.')

    # Выбираем данные из реестр и переносим в ведомости
    def runing_2(self):
        self.ui.label_info_5.clear()
        column_start = self.ui.spinBox_column_start_15.value()
        column_start_2 = self.ui.spinBox_column_start_16.value()
        row_start = self.ui.spinBox_row_start_8.value()
        row_finish = self.ui.spinBox_row_finish_8.value()
        name_book = self.ui.lineEdit_book_6.text()

        if column_start == 0:
            self.ui.label_info_5.setText('Введите № начальной колонки для № счетов.')
        if column_start_2 == 0:
            self.ui.label_info_5.setText('Введите № начальной колонки для сумм.')
        elif row_start == 0:
            self.ui.label_info_5.setText('Введите № начальной строки для расчёта.')
        elif row_finish == 0:
            self.ui.label_info_5.setText('Введите № конечной строки для расчёта.')
        elif name_book == '':
            self.ui.label_info_5.setText('Введите имя книги в ведомости.')
        else:
            try:
                wb = xl.load_workbook(self.file_reestr_1, keep_vba=True, data_only=True)
                wb.active = wb[name_book]
                ws = wb.active
                # # Выбираем все № счетов
                score_all = []
                summ_all = []
                for col_cells in ws.iter_cols(min_row=row_start, max_row=row_finish, min_col=column_start,
                                                     max_col=column_start):
                    for cell in col_cells:
                        score_all.append(cell.value)
                        # print('%s: cell.value=%s' % (cell, cell.value))
                # print('Проверка 1: ', score_all)
                for col_cells in ws.iter_cols(min_row=row_start, max_row=row_finish, min_col=column_start_2,
                                                     max_col=column_start_2):
                    for cell in col_cells:
                        ch = cell.value
                        # zam = ch.replace(".", ",")
                        if ch == None:
                            summ_all.append(ch)
                        else:
                            summ_all.append(float(ch))
                        # print('%s: cell.value=%s' % (cell, cell.value))
                # print('Проверка 1: ', summ_all)
                i = 0
                new_bill = []
                for score in score_all:
                    new_bill.append([score, summ_all[i]])
                    i+=1
            #
                # print(new_bill)

                # Находим дубликаты в списке
                dup = [x for i, x in enumerate(score_all) if i != score_all.index(x) and len(score_all[i])>15]
                # print(dup)  # [1, 5, 1]
            #
                dup_new = []
                for i in dup:
                   dup_new.append([i,777])

                # print(new_bill)
                for t in dup_new:
                    for j in new_bill:
                        if j[0] == t[0] and t[1] == 777:
                            t[1] = j[1]
                        elif j[0] == t[0] and t[1] != 777:
                            t[1] = round(t[1] + j[1], 2)

                for k in dup:
                    for j in new_bill:
                        if k == j[0]:
                            new_bill.pop(new_bill.index(j))
                for k in dup:
                    for j in new_bill:
                        if k == j[0]:
                            new_bill.pop(new_bill.index(j))

                for k in dup_new:
                    for s in new_bill:
                        if k[0] == s[0]:
                            pass
                        else:
                            new_bill.append(k)
                        break
                wb.close()
                self.ui.label_info_5.setText('Данные успешно внесены!')
            except:
                self.ui.label_info_5.setText('Проверте веденные данные по ведомости.')

            self.ui.label_info_3.clear()
            column_start_ved = self.ui.spinBox_column_start_7.value()
            column_start_ved_2 = self.ui.spinBox_column_start_8.value()
            row_start_ved = self.ui.spinBox_row_start_4.value()
            row_finish_ved = self.ui.spinBox_row_finish_4.value()
            name_book_ved = self.ui.lineEdit_book_5.text()

            if column_start_ved == 0:
                self.ui.label_info_3.setText('Введите № начальной колонки для № счетов.')
            if column_start_ved_2 == 0:
                self.ui.label_info_3.setText('Введите № начальной колонки для сумм.')
            elif row_start_ved == 0:
                self.ui.label_info_3.setText('Введите № начальной строки для расчёта.')
            elif row_finish_ved == 0:
                self.ui.label_info_3.setText('Введите № конечной строки для расчёта.')
            elif name_book_ved == '':
                self.ui.label_info_3.setText('Введите имя книги в ведомости.')
            else:
                try:
                    wb1 = xl.load_workbook(self.file_ved)
                    # print(wb1.sheetnames)
                    wb1.active = wb1[name_book_ved]
                    wb1.guess_types = True
                    ws2 = wb1.active
                    # # Выбираем все № счетов

                    score = []
                    coord = []

                    for col_cells in ws2.iter_cols(min_row=row_start_ved, max_row=row_finish_ved,
                                                   min_col=column_start_ved, max_col=column_start_ved):
                        for cell in col_cells:
                            score.append([cell.value])

                    for col_cells in ws2.iter_cols(min_row=row_start_ved, max_row=row_finish_ved,
                                                           min_col=column_start_ved_2, max_col=column_start_ved_2):
                        for cell in col_cells:
                            coord.append(cell.coordinate)
            #
                    reestr1 = []
                    t = 0
                    for i in score:
                        reestr1.append([i[0], coord[t]])
                        t += 1

                    # print(reestr1)

                    for re in reestr1:
                        for d in new_bill:
                            if re[0] == d[0]:
                                # print(d[1])
                                # print(re)
                                re.append(d[1])
                    # print(reestr1)
                    for i in reestr1:
                        try:
                            ws2[i[1]] = i[2]
                        except IndexError:
                            pass

                    wb1.save(self.file_ved)
                    self.ui.label_info_3.setText('Данные успешно внесены!')
                except:
                    self.ui.label_info_3.setText('Проверте веденные данные по ведомости.')


if __name__ == '__main__':
    import sys  # sys нужен для передачи argv в QApplication
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = DengyApp()  # Создаём объект класса Activation
    myStyle = MyProxyStyle('Fusion')
    app.setStyle(myStyle)
    window.show()  # Показываем окно активации
    sys.exit(app.exec())  # и запускаем приложение
